"""
BTRC IMS — Master Data Inventory (single sheet)
Columns: S/N, Operator, Dataset, Primary File(s), Format, Date Received,
         Received From, Records, Geometry, Sub-types, Key Attributes,
         Coord System, Data Period, Used on Map?, Gap Analysis, Notes
"""

import pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = pathlib.Path(r"C:\Users\ASUS\Downloads\BTRC-IMS-RFP-Submission-master\BTRC-IMS-RFP-Submission-master\Demo-UI\MapData\BTRC_Master_Data_Inventory.xlsx")

wb = Workbook()

# ── Palette ──────────────────────────────────────────────────────────
P = dict(
    navy="1E3A5F", blue="2563EB", blue_lt="EFF6FF", blue_acc="BFDBFE",
    green="166534", green_lt="F0FDF4", green_acc="BBF7D0",
    amber="92400E", amber_lt="FFFBEB", amber_acc="FDE68A",
    red="991B1B",  red_lt="FEF2F2",  red_acc="FECACA",
    purple="5B21B6", purple_lt="F5F3FF", purple_acc="DDD6FE",
    slate="334155", slate_lt="F8FAFC", slate_acc="E2E8F0",
    white="FFFFFF", gray="94A3B8", border="CBD5E1",
)

def sd(c=P["border"], s="thin"): return Side(border_style=s, color=c)
THIN = Border(left=sd(), right=sd(), top=sd(), bottom=sd())
MED  = Border(left=sd(P["navy"], "medium"), right=sd(P["navy"], "medium"),
              top=sd(P["navy"], "medium"),  bottom=sd(P["navy"], "medium"))

def fill(h):  return PatternFill("solid", fgColor=h)

def wcell(c, val, bold=False, sz=9, fg="000000", bg=None,
          align="left", wrap=True, italic=False, border=THIN):
    c.value = val
    c.font  = Font(name="Calibri", bold=bold, size=sz, color=fg, italic=italic)
    if bg: c.fill = fill(bg)
    c.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    c.border = border

def title_row(ws, row, text, ncols, bg=P["navy"], fg=P["white"], sz=13, h=28):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1)
    wcell(c, text, bold=True, sz=sz, fg=fg, bg=bg, align="center", border=MED)
    ws.row_dimensions[row].height = h

def sub_row(ws, row, text, ncols, bg=P["blue"], fg=P["white"], sz=10, h=16):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1)
    wcell(c, text, bold=False, sz=sz, fg=fg, bg=bg, align="center",
          italic=True, border=MED)
    ws.row_dimensions[row].height = h

def grp_row(ws, row, text, ncols, h=15):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row}")
    c = ws.cell(row=row, column=1)
    wcell(c, text, bold=True, sz=10, fg=P["navy"], bg=P["blue_acc"],
          align="left", border=THIN)
    ws.row_dimensions[row].height = h

def hdr_row(ws, row, cols, h=34):
    for ci, (label, _) in enumerate(cols, 1):
        c = ws.cell(row=row, column=ci)
        wcell(c, label, bold=True, sz=10, fg=P["white"], bg=P["blue"],
              align="center", border=THIN)
    ws.row_dimensions[row].height = h

def set_widths(ws, cols):
    for ci, (_, w) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

# ════════════════════════════════════════════════════════════════════
# MASTER SHEET
# ════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Master Data Inventory"

COLS = [
    ("S/N",                    4),
    ("Operator / Source",     22),
    ("Dataset",               22),
    ("Primary File(s)",       40),
    ("Format",                13),
    ("Date Received",         14),
    ("Received From",         20),
    ("Records / Features",    14),
    ("Geometry",              10),
    ("Sub-types / Categories",38),
    ("Key Attributes",        36),
    ("Coord System",          12),
    ("Data Period",           11),
    ("Used on Map?",          12),
    ("Gap Analysis",          40),
    ("Notes / Remarks",       40),
]
NC = len(COLS)

title_row(ws, 1,
    "BTRC IMS — Master Data Inventory  |  Submitted to BTRC for Validation",
    NC)
sub_row(ws, 2,
    "One row per dataset. Covers: received file details, source, data types,"
    " usage status, gap analysis, and remarks.",
    NC)
ws.row_dimensions[3].height = 4
hdr_row(ws, 4, COLS)
set_widths(ws, COLS)
ws.freeze_panes = "A5"

USED_MAP = {
    "Yes":      (P["green_lt"],  P["green"]),
    "Partial":  (P["amber_lt"],  P["amber"]),
    "No":       (P["red_lt"],    P["red"]),
    "Indirect": (P["purple_lt"], P["purple"]),
}

GAP_NONE    = "No gap — all records implemented."
GAP_UNKNOWN = "Not yet extracted. Content and size unknown."

# ── Master rows ──────────────────────────────────────────────────────
# Fields:
#   operator, dataset, primary_file, fmt, date_received, received_from,
#   records, geometry, subtypes, key_attrs, crs, period, used,
#   gap_analysis, notes
ROWS = [

    # ── Grameenphone ─────────────────────────────────────────────────
    ("Grameenphone (GP)",
     "BTS Sites",
     "GP_Tx Information_18-Dec-25.xlsx",
     "XLSX", "18 Dec 2025", "Grameenphone (GP)",
     "23,864", "Point",
     "Fiber Connected: 8,971\nMW (Microwave): 14,893",
     "Site Code, Latitude, Longitude, TX Type (Fiber / MW)",
     "WGS84", "Dec 2025", "Yes",
     GAP_NONE,
     "Admin (Division / District / Upazila) enriched via point-in-polygon"
     " lookup. 246 border-edge sites show '--' for admin fields."),

    # ── Robi Axiata ──────────────────────────────────────────────────
    ("Robi Axiata",
     "BTS Sites",
     "Robi site list_Nov'25_BTRC_MW & fiber.xlsx",
     "XLSX", "Nov 2025", "Robi Axiata",
     "18,838", "Point",
     "MW Only: 13,760\nFiber Only: 1,632\nMW + Fiber (Both): 3,446",
     "Site Code, Latitude, Longitude, MW flag (Yes / --), Fiber flag (Yes / --)",
     "WGS84", "Nov 2025", "Yes",
     GAP_NONE,
     "Three-category backhaul classification applied. Admin enriched via"
     " PIP. 169 border-edge sites show '--'."),

    # ── Banglalink ───────────────────────────────────────────────────
    ("Banglalink",
     "BTS Sites — Latest",
     "BTS Information_Banglalink_Till 15 Dec 25.xlsx",
     "XLSX", "Dec 2025", "Banglalink",
     "15,154", "Point",
     "Microwave: 13,015\nFiber: 2,113\nInactive: 26",
     "Unique Site ID, Lat, Lon, Division, District, Thana, Address,"
     " MW Link Status, Fiber Optic Status, Primary Backhaul Type",
     "WGS84", "Dec 2025", "Yes",
     GAP_NONE,
     "Admin data taken directly from Excel (no PIP needed)."
     " Address field included in popup."),

    ("Banglalink",
     "BTS Towers — Historical",
     "3gtower.geojson",
     "GeoJSON", "Pre-2025", "Banglalink",
     "13,208", "Point",
     "4G sites\n3G only\n2G only",
     "Site Code, Site Name, Generation (2G/3G/4G), Division, District,"
     " Upazila, Union, Vendor",
     "WGS84", "Pre-2025", "Yes",
     GAP_NONE,
     "Historical tower dataset. Filterable by generation on map."),

    ("Banglalink",
     "Fiber Lines — Historical",
     "bl-line.geojson",
     "GeoJSON", "Pre-2025", "Banglalink",
     "172", "LineString",
     "72 Core\n48 Core\n32 Core",
     "Core Count",
     "WGS84", "Pre-2025", "Yes",
     GAP_NONE,
     "Historical fiber route lines. Filterable by core count on map."),

    # ── BTCL ─────────────────────────────────────────────────────────
    ("BTCL",
     "Network Points — Latest",
     "GEO SPIRAL DATA STRUCTURE_TEMPLE_FINAL_BTCL.xlsx",
     "XLSX", "2025", "BTCL",
     "24,142", "Point",
     "CP (Connection Point)\nHH (Hand Hole)\nHOP\nPOP\nMH (Man Hole)\nOther"
     "\n(3,432 raw type values normalised to 6 categories)",
     "Name, Latitude, Longitude, Point Type, Raw Type, Year, Feature Code",
     "WGS84", "2025", "Yes",
     "69 records removed (0.3% of total)."
     " Coordinates fell outside Bangladesh boundary — likely data-entry errors.",
     "24,073 of 24,142 on map. Filtered points near borders should be"
     " verified and resubmitted by BTCL with corrected coordinates."),

    ("BTCL",
     "Fiber Lines — Historical",
     "btcl-nttn-line.geojson",
     "GeoJSON", "Pre-2025", "BTCL",
     "584", "MultiLineString",
     "144+ Core\n96 Core\n48 Core\n24 Core\n<24 Core",
     "Operator Name, Line Name, Line Type, Core Count, Route Length (km),"
     " Cable Length (km), Duct Info, Division / District / Upazila",
     "WGS84", "Pre-2025", "Yes",
     GAP_NONE,
     "All lines on map. Filterable by core-count band."),

    ("BTCL",
     "Network Nodes — Historical",
     "btcl-ponts.geojson",
     "GeoJSON", "Pre-2025", "BTCL",
     "29,795", "Point",
     "HOP\nHH (Hand Hole)\nCP (Connection Point)\nMH (Man Hole)",
     "Operator Name, Point Name, Point Type, Lat, Lon,"
     " Division, District, Upazila, Union, Mouza, Year",
     "WGS84", "Pre-2025", "Yes",
     GAP_NONE,
     "All nodes on map. Filterable by node type."),

    ("BTCL",
     "Union Project Locations",
     "btcl-union-project-location.geojson",
     "GeoJSON", "Pre-2025", "BTCL",
     "966", "Point",
     "Union-level project markers",
     "Location coordinates",
     "WGS84", "Pre-2025", "Yes",
     GAP_NONE,
     "All union project markers on map."),

    # ── Fiber@Home FHLFON ─────────────────────────────────────────────
    ("Fiber@Home (FHLFON)",
     "Fiber Lines",
     "FHLFONLine.shp + FHLFONLineExcel.xlsx\n"
     "(9 files: 8 shapefile components + 1 Excel)",
     "Shapefile + XLSX", "Dec 2024", "Fiber@Home / FHLFON",
     "141,567", "LineString",
     "Aerial: 85,366\nBurial: 53,870\nOPGW: 2,331",
     "Operator Code, Operator Name, Infra Type (Own / Leased), Line ID,"
     " Year, Month, Line Name, Line Type, Path Along, Duct No, Duct Use, Cable No",
     "WGS84", "Dec 2024", "Partial",
     "123,162 lines not yet on map (87% gap).\n"
     "Currently showing: 18,405 of 141,567.\n"
     "Missing: full Aerial set (85k), full Burial set (53k), all OPGW (2,331).",
     "Full re-ingestion from FHLFONLineExcel.xlsx required."
     " Current map uses only a pre-converted subset."),

    ("Fiber@Home (FHLFON)",
     "Network Points",
     "FHLFONPoint.shp + FHLFONPointExcel.xlsx\n"
     "(10 files: 8 shapefile + 1 Excel + 1 duplicate)",
     "Shapefile + XLSX", "Dec 2024", "Fiber@Home / FHLFON",
     "122,150", "Point",
     "JE (Joint Enclosure): 38,097\nEP (End Point): 25,804\nCoupler: 22,435"
     "\nFL (Fiber Loop): 16,558\nHH (Hand Hole): 8,292\nBTS: 4,755"
     "\nCO (Central Office): 2,087\nFAT: 1,591\nBDB: 1,390\nPole: 479"
     "\nLDP: 387\nFDH: 80\nPIT: 26\nCustomer: 25\nOther / Noise: ~44",
     "Year, Month, Operator Code, Operator Name, Point ID, Point Type,"
     " Point Name, Address, Latitude, Longitude, Feature Code",
     "WGS84", "Dec 2024", "Partial",
     "27,197 points not on map (22% gap).\n"
     "Currently showing: 94,953 of 122,150.\n"
     "Missing types: Coupler (22,435), FL (16,558), BDB (1,390),"
     " Pole (479), LDP (387), PIT (26), Customer (25).",
     "'Copy of FHLFONPointExcel.xlsx' is an exact duplicate and not used."),

    # ── Summit Communications ─────────────────────────────────────────
    ("Summit Communications",
     "Fiber Lines",
     "Line_data.xlsx + Line_data.shp\n(1 Excel + multiple shapefile components)",
     "XLSX + Shapefile", "Mar 2026", "Summit Communications",
     "102,514", "LineString",
     "Aerial: 78,499\nBurial: 22,702\nOverhead PGCB: 1,209"
     "\nBridge Crossing: 14\nBurial (Damaged): 89",
     "FID, Line ID, Year, Month, Operator Code, Operator Name, Infra Type,"
     " Line Type, Line Name, Path Along, No of Ducts, Duct Use",
     "WGS84", "Mar 2026", "Partial",
     "79,357 lines not on map (77% gap).\n"
     "Currently showing: 23,157 of 102,514.\n"
     "Missing: majority of Aerial (78k) and Burial (22k) lines.",
     "Full re-ingestion from Line_data.xlsx required. Current map uses a"
     " simplified 3-tier classification only."),

    ("Summit Communications",
     "Network Points",
     "Point_data.xlsx + Point_data.shp\n(1 Excel + multiple shapefile components)",
     "XLSX + Shapefile", "Mar 2026", "Summit Communications",
     "68,850", "Point",
     "TJB (Joint Box): 37,909\nHH (Hand Hole): 9,506\nBTS: 8,613"
     "\nEP (End Point): 6,872\nNode / Info: 1,539\nODB: 1,193\nNode / TT: 812"
     "\nPOP: 666\nPOC: 453\nBS (Base Station): 422\nNode / CBD: 245"
     "\nCOLO / PoP: 128\nFAT: 89\nFDT: 62\nOther: ~183",
     "FID, Point ID, Year, Month, Operator Code, Operator Name,"
     " Point Type, Point Name, Latitude, Longitude, Feature Code",
     "WGS84", "Mar 2026", "Partial",
     "54,288 points not on map (79% gap).\n"
     "Currently showing: 14,562 of 68,850.\n"
     "Missing: TJB (37,909), EP (6,872), ODB (1,193), Node/TT (812),"
     " POP (666), POC (453), BS (422), Node/CBD (245), COLO/PoP (128), FAT (89), FDT (62).",
     "Only Nodes (5,949) and BTS (8,613) currently on map."),

    ("Summit Communications",
     "Summit.rar (Archive — not extracted)",
     "Summit.rar",
     "RAR Archive", "Not specified", "Summit Communications",
     "Unknown", "Unknown",
     "Unknown — archive not extracted",
     "Unknown",
     "Unknown", "Unknown", "No",
     GAP_UNKNOWN,
     "Archive file found in Railway folder. Must be extracted and analysed"
     " to determine if it contains additional network data."),

    # ── Bahon Limited ─────────────────────────────────────────────────
    ("Bahon Limited",
     "Fiber Lines",
     "Bahon Network_System Line.shp\n"
     "(+ .dbf, .shx, .sbn, .sbx, .cpg, .prj, .qmd)",
     "Shapefile", "Pre-2025", "Bahon Limited",
     "~7,763", "LineString",
     "Overhead (OH)\nUnderground (UG)\nWall Clamped (WC)",
     "Cable Type, Division, District, Upazila (from DBF attributes)",
     "WGS84", "Pre-2025", "Yes",
     "No significant gap detected. Exact source count unverified"
     " (shapefile only — no Excel counterpart for cross-check).",
     "No Excel submission received from Bahon. Request Excel tabular"
     " data for BTRC records."),

    ("Bahon Limited",
     "Network Nodes",
     "(Derived from shapefile dataset)",
     "Shapefile", "Pre-2025", "Bahon Limited",
     "12,817", "Point",
     "Network junction nodes",
     "Node coordinates",
     "WGS84", "Pre-2025", "Yes",
     GAP_NONE,
     "All nodes on map."),

    # ── InfoSarkar-3 ──────────────────────────────────────────────────
    ("InfoSarkar-3 (IS3)",
     "Fiber Lines",
     "doc.kml\n"
     "(icon assets: mysavedplaces_closed.png, mysavedplaces_open.png)",
     "KML", "Pre-2025", "InfoSarkar-3 / FHL",
     "3,383", "LineString",
     "48 Core\n24 Core\n12 Core\nMessenger\nRing\nCBD",
     "Core count, Name, Layer, Length (km)",
     "WGS84", "Pre-2025", "Yes",
     GAP_NONE,
     "All lines on map with core-count filter."),

    ("InfoSarkar-3 (IS3)",
     "Network Nodes",
     "doc.kml (same file)",
     "KML", "Pre-2025", "InfoSarkar-3 / FHL",
     "477", "Point",
     "Network junction nodes",
     "Node coordinates",
     "WGS84", "Pre-2025", "Yes",
     GAP_NONE,
     "All nodes on map."),

    # ── PGCB ──────────────────────────────────────────────────────────
    ("PGCB",
     "OPGW Transmission Lines",
     "Power Grid Tranmission(OPGW).kml",
     "KML", "Pre-2025", "PGCB",
     "324", "LineString",
     "400 kV Transmission Line\n230 kV Transmission Line"
     "\n132 kV Transmission Line\nUnderground Cable\nOthers",
     "Layer (voltage / type), Name, Description",
     "WGS84", "Pre-2025", "Yes",
     GAP_NONE,
     "All OPGW lines on map. Filterable by voltage / line type."),

    # ── Multi-operator Fiber Network ──────────────────────────────────
    ("Multi-operator\n(GP, Robi, BTCL, BL, MOTN, BSCCL)",
     "Fiber Network — Lines",
     "fiber_network_multiple_district.kmz",
     "KMZ", "2025", "BTRC / Multiple Operators",
     "8,163", "LineString",
     "Grameenphone\nRobi\nBTCL\nBanglalink\nMOTN\nBSCCL\nUnknown",
     "Name, Operator, Distance (km)",
     "WGS84", "2025", "Yes",
     GAP_NONE,
     "All 6-operator inter-district fiber lines on map with operator colour filter."),

    ("Multi-operator\n(GP, Robi, BTCL, BL, MOTN, BSCCL)",
     "Fiber Network — Points",
     "fiber_network_multiple_district.kmz (same file)",
     "KMZ", "2025", "BTRC / Multiple Operators",
     "19,096", "Point",
     "Grameenphone\nRobi\nBTCL\nBanglalink\nMOTN\nBSCCL\nUnknown",
     "Name, Operator",
     "WGS84", "2025", "Yes",
     GAP_NONE,
     "All 6-operator fiber junction points on map."),

    # ── Bangladesh Railway ────────────────────────────────────────────
    ("Bangladesh Railway (BR)",
     "Fiber Lines along Railway",
     "Geo Spatial Data Structure_Template_Final_railway.xlsx",
     "XLSX", "2025", "Bangladesh Railway",
     "353", "LineString",
     "8 Core\n16 Core\n32 Core\n48 Core\n72 Core\n96 Core",
     "Station Name A, Station Name B, Length (km),"
     " Total Core, Used Core, Unused Core",
     "WGS84", "2025", "Yes",
     GAP_NONE,
     "All 353 fiber route segments on map. Duplicate copy of same file"
     " exists in Railway folder — not used."),

    ("Bangladesh Railway (BR)",
     "Fiber Station Nodes",
     "(Derived from railway Excel)",
     "XLSX", "2025", "Bangladesh Railway",
     "354", "Point",
     "Station junction nodes",
     "Station name",
     "WGS84", "2025", "Yes",
     GAP_NONE,
     "All 354 station nodes on map."),

    ("Bangladesh Railway (BR)",
     "Railway Track (Base Layer)",
     "railway.geojson\n"
     "(source: railline_wgs.shp + .dbf + .shx + .prj + .cst)",
     "GeoJSON + Shapefile", "Pre-2025", "Bangladesh Railway",
     "2,675", "LineString",
     "Railway route segments",
     "Route geometry",
     "WGS84", "Pre-2025", "Yes",
     GAP_NONE,
     "Base railway track layer on map."),

    # ── ISP POP ───────────────────────────────────────────────────────
    ("All ISPs (Multi-operator)",
     "ISP POP Locations",
     "aisp-pop.geojson  /  all-isp-pop-info.json\n(same data in two formats)",
     "GeoJSON + JSON", "2020-2021", "BTRC / All ISPs",
     "3,930", "Point",
     "Category-A ISPs\nCategory-B ISPs\nCategory-C ISPs"
     "\n(covers all districts nationwide)",
     "GID, POP Code, POP Name, POP Address, POP Capacity, ISP Name,"
     " Type of ISP, Lat, Lon, District, Division, Upazila, Union, Mouza",
     "WGS84", "2020-2021", "No",
     "3,930 records not yet on map (100% gap).\n"
     "Data is fully available and ready for implementation.",
     "High priority. Add ISP POP overlay with category filter (A/B/C)"
     " and district filter. Popup should show ISP name, capacity, address, admin."),
]

# ── Render rows ──────────────────────────────────────────────────────
row  = 5
alt  = False
prev = None
sn   = 0

for (op, dataset, pfile, fmt, date_recv, recv_from,
     records, geom, subtypes, attrs, crs, period, used,
     gap, notes) in ROWS:

    if op != prev:
        grp_row(ws, row, f"  {op}", NC)
        row += 1
        prev = op
        alt  = False

    sn += 1
    used_key = used.split(" ")[0]
    bg_used, fg_used = USED_MAP.get(used_key, (P["white"], P["slate"]))
    row_bg = P["slate_lt"] if alt else P["white"]

    # gap analysis bg
    gap_bg = (P["amber_lt"] if "gap" in gap.lower() and "No gap" not in gap
              else (P["red_lt"] if "100% gap" in gap
              else (P["green_lt"] if gap == "No gap — all records implemented."
              else row_bg)))
    gap_fg = (P["amber"]   if "gap" in gap.lower() and "No gap" not in gap
              else (P["red"]   if "100% gap" in gap
              else (P["green"] if gap == "No gap — all records implemented."
              else P["slate"])))

    vals = [sn, op, dataset, pfile, fmt, date_recv, recv_from,
            records, geom, subtypes, attrs, crs, period, used, gap, notes]

    for ci, val in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci)
        if ci == 14:   # Used on Map
            wcell(c, val, bold=True, sz=9, fg=fg_used, bg=bg_used, align="center")
        elif ci == 15:  # Gap Analysis
            wcell(c, val, sz=9, fg=gap_fg, bg=gap_bg)
        elif ci == 1:
            wcell(c, val, bold=True, sz=9, bg=row_bg, align="center")
        elif ci == 2:
            wcell(c, val, bold=True, sz=9, bg=row_bg)
        else:
            wcell(c, val, sz=9, bg=row_bg)

    ws.row_dimensions[row].height = 80
    alt = not alt
    row += 1

# ── Legend ───────────────────────────────────────────────────────────
row += 1
for label, (bg, fg) in [
    ("Yes — fully implemented on the map",           (P["green_lt"],  P["green"])),
    ("Partial — only a subset is on the map",        (P["amber_lt"],  P["amber"])),
    ("No — data not yet implemented on the map",     (P["red_lt"],    P["red"])),
    ("Indirect — used as a source for another file", (P["purple_lt"], P["purple"])),
]:
    ws.merge_cells(f"A{row}:{get_column_letter(NC)}{row}")
    c = ws.cell(row=row, column=1)
    wcell(c, f"  [legend]  {label}", bold=True, sz=9, fg=fg, bg=bg, border=THIN)
    ws.row_dimensions[row].height = 14
    row += 1

# ── Sheet settings ───────────────────────────────────────────────────
ws.auto_filter.ref = f"A4:{get_column_letter(NC)}4"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToPage   = True
ws.page_setup.fitToWidth  = 1
ws.page_setup.fitToHeight = 0
ws.print_title_rows       = "1:4"
ws.sheet_view.showGridLines = True

wb.save(str(OUT))
print(f"Saved  -> {OUT}")
print(f"Rows   : {sn} datasets")
print(f"Columns: {NC}")
