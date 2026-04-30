"""
Step 1: Rebuild fhlfon-points.geojson from FHLFONPointExcel.xlsx
- Normalize dirty point_type values
- Filter to CO, BTS, JE, EP, FAT, FDH
- Parse district from address
- Output compact GeoJSON with short property keys
"""
import json
import re
import openpyxl

SRC = "FHLFON/FHLFONPointExcel.xlsx"
OUT_PRIMARY = "../public/data/fhlfon-points-primary.geojson"   # CO, BTS, FDH — always loaded
OUT_DETAIL  = "../public/data/fhlfon-points-detail.geojson"    # JE, EP, FAT  — zoom ≥ 11 only

PRIMARY_TYPES = {"CO", "BTS", "FDH"}
DETAIL_TYPES  = {"JE", "EP", "FAT"}

# Keep only these normalized types
KEEP = PRIMARY_TYPES | DETAIL_TYPES

# Regex-based normalizer — order matters (most specific first)
NORM_RULES = [
    (re.compile(r"^CO",        re.I), "CO"),
    (re.compile(r"^BTS",       re.I), "BTS"),
    (re.compile(r"^FAT",       re.I), "FAT"),
    (re.compile(r"^FDH",       re.I), "FDH"),
    (re.compile(r"^JE",        re.I), "JE"),
    (re.compile(r"^EP",        re.I), "EP"),
    (re.compile(r"^Coupler",   re.I), "Coupler"),
    (re.compile(r"^FL",        re.I), "FL"),
    (re.compile(r"^HandHole",  re.I), "HandHole"),
    (re.compile(r"^BDB",       re.I), "BDB"),
]

def normalize_type(raw):
    if not raw:
        return None
    raw = str(raw).strip()
    for pattern, label in NORM_RULES:
        if pattern.match(raw):
            return label
    return None

def parse_district(address):
    if not address:
        return ""
    parts = str(address).split(",")
    # Last non-empty token, stripped
    for token in reversed(parts):
        t = token.strip()
        if t:
            return t
    return ""

print(f"Loading {SRC} …")
wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
ws = wb.active

# Read header row
headers = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
print(f"Columns: {headers}")

# Map column names → indices (0-based)
col = {h: i for i, h in enumerate(headers)}

# We need: point_type, point_name, latitude, longitude, address
needed = ["point_type", "point_name", "latitude", "longitude", "address"]
for n in needed:
    if n not in col:
        # case-insensitive fallback
        matches = [h for h in col if h.lower() == n.lower()]
        if matches:
            col[n] = col[matches[0]]
        else:
            print(f"WARNING: column '{n}' not found. Available: {headers}")

primary_features = []
detail_features  = []
skipped = 0
kept_counts = {}

for row in ws.iter_rows(min_row=2, values_only=True):
    raw_type = row[col["point_type"]] if "point_type" in col else None
    pt = normalize_type(raw_type)

    if pt not in KEEP:
        skipped += 1
        continue

    try:
        lat = float(row[col["latitude"]])
        lon = float(row[col["longitude"]])
    except (TypeError, ValueError):
        skipped += 1
        continue

    # Skip zero/null coordinates
    if lat == 0.0 and lon == 0.0:
        skipped += 1
        continue

    nm  = str(row[col["point_name"]] or "").strip() if "point_name" in col else ""
    adr = row[col["address"]] if "address" in col else ""
    dist = parse_district(adr)

    feat = {
        "type": "Feature",
        "geometry": {"type": "Point", "coordinates": [round(lon, 6), round(lat, 6)]},
        "properties": {"pt": pt, "nm": nm, "dist": dist},
    }
    if pt in PRIMARY_TYPES:
        primary_features.append(feat)
    else:
        detail_features.append(feat)
    kept_counts[pt] = kept_counts.get(pt, 0) + 1

wb.close()

import os

def write_geojson(path, features):
    gj = {"type": "FeatureCollection", "features": features}
    with open(path, "w", encoding="utf-8") as f:
        json.dump(gj, f, separators=(",", ":"), ensure_ascii=False)
    return os.path.getsize(path) / 1024

sz_p = write_geojson(OUT_PRIMARY, primary_features)
sz_d = write_geojson(OUT_DETAIL,  detail_features)

total = len(primary_features) + len(detail_features)
print(f"\nDone. {total} points kept, {skipped} skipped.")
print(f"Type breakdown: {kept_counts}")
print(f"Primary ({', '.join(sorted(PRIMARY_TYPES))}): {len(primary_features)} pts -> {OUT_PRIMARY}  ({sz_p:.1f} KB)")
print(f"Detail  ({', '.join(sorted(DETAIL_TYPES))}):  {len(detail_features)} pts -> {OUT_DETAIL}   ({sz_d:.1f} KB)")
