"""
Generate BTRC Data Inventory Excel for validation submission.
"""

import pathlib
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

OUT = pathlib.Path(r"C:\Users\ASUS\Downloads\BTRC-IMS-RFP-Submission-master\BTRC-IMS-RFP-Submission-master\Demo-UI\MapData\BTRC_Data_Inventory.xlsx")

wb = Workbook()
ws = wb.active
ws.title = "Data Inventory"

# ── Colour palette ───────────────────────────────────────────────────
C_TITLE_BG   = "1E3A5F"   # dark navy
C_TITLE_FG   = "FFFFFF"
C_HDR_BG     = "2563EB"   # blue header
C_HDR_FG     = "FFFFFF"
C_GRP1_BG    = "EFF6FF"   # light blue  – latest submissions
C_GRP2_BG    = "F0FDF4"   # light green – existing operator data
C_GRP1_ACC   = "BFDBFE"   # group header accent
C_GRP2_ACC   = "BBF7D0"
C_ALT        = "F8FAFC"   # alternating row tint
C_WHITE      = "FFFFFF"
C_BORDER     = "CBD5E1"

def side(color=C_BORDER, style="thin"):
    return Side(border_style=style, color=color)

THIN  = Border(left=side(), right=side(), top=side(), bottom=side())
THICK = Border(left=side("1E3A5F","medium"), right=side("1E3A5F","medium"),
               top=side("1E3A5F","medium"), bottom=side("1E3A5F","medium"))

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def hdr_font(bold=True, size=10, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def cell_font(bold=False, size=9, color="000000"):
    return Font(name="Calibri", bold=bold, size=size, color=color)

def wrap(ws_cell, value, font=None, fill_color=None, align="left", bold=False, size=9):
    ws_cell.value = value
    ws_cell.font  = font or Font(name="Calibri", bold=bold, size=size)
    if fill_color:
        ws_cell.fill = fill(fill_color)
    ws_cell.alignment = Alignment(
        horizontal=align, vertical="top", wrap_text=True
    )
    ws_cell.border = THIN

# ── Column definitions ────────────────────────────────────────────────
COLUMNS = [
    ("S/N",              4),
    ("Operator / Source", 18),
    ("Source File Name",  36),
    ("File Format",        9),
    ("Date Received",     13),
    ("Data Type",         18),
    ("Geometry Type",     13),
    ("Total Records",     13),
    ("Key Fields / Attributes", 36),
    ("Categories / Backhaul Types", 30),
    ("Geographic Coverage", 16),
    ("Map Layer Name",    22),
    ("Implementation Status", 14),
    ("Remarks",           34),
]

# ── Title row ─────────────────────────────────────────────────────────
ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
title_cell = ws["A1"]
title_cell.value   = "BTRC IMS — Received Data Inventory & Validation Sheet"
title_cell.font    = Font(name="Calibri", bold=True, size=14, color=C_TITLE_FG)
title_cell.fill    = fill(C_TITLE_BG)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 28

# Sub-title row
ws.merge_cells(f"A2:{get_column_letter(len(COLUMNS))}2")
sub = ws["A2"]
sub.value     = "Prepared for BTRC Validation  |  Reference: BTRC IMS RFP Submission  |  Date: May 2026"
sub.font      = Font(name="Calibri", italic=True, size=10, color="FFFFFF")
sub.fill      = fill("2563EB")
sub.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 18

# Blank spacer
ws.row_dimensions[3].height = 6

# ── Column headers ────────────────────────────────────────────────────
for col_idx, (label, _) in enumerate(COLUMNS, start=1):
    c = ws.cell(row=4, column=col_idx)
    c.value = label
    c.font  = Font(name="Calibri", bold=True, size=10, color=C_HDR_FG)
    c.fill  = fill(C_HDR_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = THIN
ws.row_dimensions[4].height = 32

# Set column widths
for col_idx, (_, width) in enumerate(COLUMNS, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── Data rows ──────────────────────────────────────────────────────────
# fmt: (sn, operator, filename, format, date, data_type, geom, records,
#        key_fields, categories, coverage, layer_name, status, remarks)

GROUP1_LABEL = "── GROUP 1: Latest BTRC Submissions (2025) ──"
GROUP2_LABEL = "── GROUP 2: Previously Integrated Operator Data ──"

ROWS = [
    # ── Group 1 header ──
    ("GROUP1",),

    (1,
     "Grameenphone (GP)",
     "GP_Tx Information_18-Dec-25.xlsx",
     "XLSX",
     "18-Dec-2025",
     "BTS Sites",
     "Point",
     "23,864",
     "Site Code, Latitude, Longitude, TX Type (Fiber / MW)",
     "Fiber Connected: 8,971\nMW Connected: 14,893",
     "Nationwide",
     "Grameenphone → Sites",
     "Implemented",
     "All coordinates valid. Admin (Division/District/Upazila) enriched via PIP against BD boundary. 246 border-edge sites unmatched.",
    ),
    (2,
     "Robi Axiata",
     "Robi site list_Nov'25_BTRC_MW & fiber.xlsx",
     "XLSX",
     "Nov-2025",
     "BTS Sites",
     "Point",
     "18,838",
     "Site Code, Latitude, Longitude, MW (Yes/–), Fiber (Yes/–)",
     "MW Only: 13,760\nFiber Only: 1,632\nMW + Fiber: 3,446",
     "Nationwide",
     "Robi → Sites",
     "Implemented",
     "Admin enriched via PIP. 169 border-edge sites unmatched. Sites with both MW and Fiber classified as 'Both'.",
    ),
    (3,
     "Banglalink",
     "BTS Information_Banglalink_Till 15 Dec 25.xlsx",
     "XLSX",
     "15-Dec-2025",
     "BTS Sites",
     "Point",
     "15,154",
     "Unique Site ID, Latitude, Longitude, Division, District, Thana, Address, MW Link Status, Fiber Optic Status, Primary Backhaul Type",
     "Microwave: 13,015\nFiber: 2,113\nInactive: 26",
     "Nationwide",
     "Banglalink → BTS Sites (Latest)",
     "Implemented",
     "Admin data (Division/District/Thana) already present in file. No PIP required. 26 inactive sites included.",
    ),
    (4,
     "BTCL",
     "GEO SPIRAL DATA STRUCTURE_TEMPLE_FINAL_BTCL.xlsx",
     "XLSX",
     "2025",
     "Network Points",
     "Point",
     "24,073\n(raw: 24,142)",
     "Name, Latitude, Longitude, Point Type, Raw Type, Year, Feature Code",
     "CP (Connection Point)\nHH (Hand Hole)\nHOP\nPOP\nMH (Man Hole)\nOther",
     "Nationwide",
     "BTCL → Points (Excel 2025)",
     "Implemented",
     "69 points removed via PIP filter (outside Bangladesh border). 3,432 unique raw type values normalised into 6 canonical categories.",
    ),
    (5,
     "Multi-Operator\n(GP, Robi, BTCL, Banglalink, MOTN, BSCCL)",
     "fiber_network_multiple_district.kmz",
     "KMZ",
     "2025",
     "Fiber Lines & Junction Points",
     "LineString / Point",
     "Lines: 8,163\nPoints: 19,096",
     "Name, Operator, Distance (km)",
     "GP, Robi, BTCL,\nBanglalink, MOTN, BSCCL,\nUnknown",
     "Nationwide",
     "Fiber Network → Lines / Points",
     "Implemented",
     "KMZ converted to GeoJSON. Operator names normalised. Lines and points rendered separately with operator-based colour coding.",
    ),
    (6,
     "Bangladesh Railway (BR)",
     "Geo Spatial Data Structure_Template_Final_railway.xlsx",
     "XLSX",
     "2025",
     "Railway Fiber Network",
     "LineString / Point",
     "Lines: 353\nStation Nodes: 354",
     "Station Name (A→B), Length (km), Total Core, Used Core, Unused Core",
     "8 / 16 / 32 / 48 / 72 / 96 Core",
     "Nationwide",
     "Railway → Bangladesh Railway Fiber",
     "Implemented",
     "Operator lines on railway right-of-way also included (Summit, Bahon, Fiber@Home, Banglalink, Robi, GP).",
    ),

    # ── Group 2 header ──
    ("GROUP2",),

    (7,
     "BTCL (Historical)",
     "btcl-nodes.geojson\nbtcl-lines.geojson",
     "GeoJSON",
     "Pre-2025",
     "Fiber Network (Historical)",
     "Point / MultiLineString",
     "Nodes: 29,795\nLines: 584",
     "Operator Name, Point Type (HOP/HH/CP/MH), Core Count, Division, District, Upazila, Union, Mouza, Year",
     "HOP, HH (Handhole),\nCP (Connection Point),\nMH (Manhole)\nCore: 144/96/48/24/<24",
     "Nationwide",
     "Telecom Operators → BTCL-OLD",
     "Implemented",
     "Older BTCL submission. Shown alongside latest Excel data. Union project locations also included (966 points).",
    ),
    (8,
     "PGCB",
     "opgw-lines.geojson",
     "GeoJSON",
     "Pre-2025",
     "OPGW Transmission Lines",
     "LineString",
     "324",
     "Layer (voltage level), Name, Description",
     "400 kV T/L\n230 kV T/L\n132 kV T/L\nUnderground Cable\nOthers",
     "Nationwide",
     "Telecom Operators → PGCB",
     "Implemented",
     "Power Grid Company of Bangladesh OPGW fiber routes along transmission lines.",
    ),
    (9,
     "Bahon Limited",
     "bahon-lines.geojson\nbahon-points.geojson",
     "GeoJSON",
     "Pre-2025",
     "Fiber Network",
     "LineString / Point",
     "Lines: 7,763\nNodes: 12,817",
     "Cable Type (OH/UG/WC), Division, District, Upazila",
     "Overhead (OH)\nUnderground (UG)\nWall Clamped (WC)\nNetwork Nodes",
     "Nationwide",
     "Telecom Operators → NTTN → Bahon",
     "Implemented",
     "NTTN operator fiber network.",
    ),
    (10,
     "InfoSarkar-3 (IS3)",
     "is3-lines.geojson\nis3-points.geojson",
     "GeoJSON",
     "Pre-2025",
     "Fiber Network",
     "LineString / Point",
     "Lines: 3,383\nNodes: 477",
     "Core Count, Name, Layer, Length (km)",
     "48 Core / 24 Core / 12 Core\nMessenger / Ring / CBD",
     "Nationwide",
     "Telecom Operators → NTTN → InfoSarkar-3",
     "Implemented",
     "Government NTTN network (InfoSarkar-3).",
    ),
    (11,
     "Fiber@Home (FHLFON)",
     "fhlfon-lines.geojson\nfhlfon-points-primary.geojson\nfhlfon-points-detail.geojson",
     "GeoJSON",
     "Pre-2025",
     "Fiber Network",
     "LineString / Point",
     "Lines: 18,405\nPoints: 94,953",
     "Line Type (Aerial/Burial), Point Type (CO/BTS/FDH/JE/EP/FAT)",
     "Aerial / Burial\nCO, BTS, FDH,\nJE, EP, FAT",
     "Nationwide",
     "Telecom Operators → NTTN → Fiber@Home",
     "Implemented",
     "Large dataset — 94,953 detail points split across two layers for performance.",
    ),
    (12,
     "Summit Communications",
     "summit-lines-backbone.geojson\nsummit-lines-major.geojson\nsummit-lines-infra.geojson\nsummit-nodes.geojson\nsummit-bts.geojson",
     "GeoJSON",
     "Pre-2025",
     "Fiber Network + BTS Sites",
     "LineString / Point",
     "Lines: 23,157\nNodes: 5,949\nBTS: 8,613",
     "Line Type (Backbone/Major/PGCB/Railway), Point Type, Point Name",
     "Backbone (≥96 core)\nMajor Burial (48–95 core)\nPGCB Route\nRailway Route\nNetwork Nodes / BTS",
     "Nationwide",
     "Telecom Operators → NTTN → Summit",
     "Implemented",
     "Summit network split into 3 line layers by importance + separate node and BTS point layers.",
    ),
    (13,
     "Banglalink (Historical Towers)",
     "bl-towers.geojson",
     "GeoJSON",
     "Pre-2025",
     "BTS Tower Sites",
     "Point",
     "13,208",
     "Site Code, Site Name, Generation (2G/3G/4G), Division, District, Upazila, Union, Vendor",
     "4G / 3G / 2G",
     "Nationwide",
     "Telecom Operators → Mobile → Banglalink → Towers",
     "Implemented",
     "Historical Banglalink tower dataset. Complement to the latest Dec-2025 BTS submission.",
    ),
    (14,
     "Banglalink (Historical Fiber)",
     "bl-lines.geojson",
     "GeoJSON",
     "Pre-2025",
     "Fiber Lines",
     "MultiLineString",
     "172",
     "Core Count",
     "72 Core / 48 Core / 32 Core",
     "Nationwide",
     "Telecom Operators → Mobile → Banglalink → Fiber Lines",
     "Implemented",
     "Historical Banglalink fiber route dataset.",
    ),
    (15,
     "Bangladesh Railway",
     "bd-railways.geojson",
     "GeoJSON",
     "Pre-2025",
     "Railway Network (Track)",
     "LineString",
     "2,675",
     "Route geometry",
     "—",
     "Nationwide",
     "Railway → Railline",
     "Implemented",
     "Base railway track layer used as reference beneath fiber overlays.",
    ),
]

# ── Write rows ─────────────────────────────────────────────────────────
current_row = 5
alt = False

for row_data in ROWS:
    # Group header row
    if len(row_data) == 1 and row_data[0].startswith("GROUP"):
        is_g1 = row_data[0] == "GROUP1"
        label = GROUP1_LABEL if is_g1 else GROUP2_LABEL
        accent = C_GRP1_ACC if is_g1 else C_GRP2_ACC
        ws.merge_cells(f"A{current_row}:{get_column_letter(len(COLUMNS))}{current_row}")
        c = ws.cell(row=current_row, column=1)
        c.value     = label
        c.font      = Font(name="Calibri", bold=True, size=10, color="1E3A5F")
        c.fill      = fill(accent)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border    = THIN
        ws.row_dimensions[current_row].height = 16
        current_row += 1
        alt = False
        continue

    # Data row
    bg = C_GRP1_BG if current_row <= 10 else C_GRP2_BG
    row_bg = C_ALT if alt else C_WHITE

    sn, operator, fname, fmt, date, dtype, geom, records, fields, cats, cov, layer, status, remarks = row_data

    status_color = {
        "Implemented": "166534",
        "Pending":     "92400E",
        "Partial":     "1D4ED8",
    }.get(status, "374151")

    row_vals = [sn, operator, fname, fmt, date, dtype, geom, records, fields, cats, cov, layer, status, remarks]

    for col_idx, val in enumerate(row_vals, start=1):
        c = ws.cell(row=current_row, column=col_idx)
        c.value  = str(val)
        c.border = THIN
        c.alignment = Alignment(horizontal="left" if col_idx > 1 else "center",
                                vertical="top", wrap_text=True)

        # Status column special colour
        if col_idx == 13:
            c.font = Font(name="Calibri", bold=True, size=9, color=status_color)
            c.fill = fill(row_bg)
        else:
            c.font = Font(name="Calibri", size=9)
            c.fill = fill(row_bg)

    # Bold the S/N and operator columns
    ws.cell(row=current_row, column=1).font = Font(name="Calibri", bold=True, size=9)
    ws.cell(row=current_row, column=2).font = Font(name="Calibri", bold=True, size=9)

    ws.row_dimensions[current_row].height = 60
    alt = not alt
    current_row += 1

# ── Summary row ───────────────────────────────────────────────────────
summary_row = current_row + 1
ws.merge_cells(f"A{summary_row}:G{summary_row}")
c = ws.cell(row=summary_row, column=1)
c.value     = "TOTAL DATASETS: 15   |   Latest Submissions: 6   |   Pre-integrated Data: 9"
c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
c.fill      = fill(C_TITLE_BG)
c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[summary_row].height = 20

ws.merge_cells(f"H{summary_row}:{get_column_letter(len(COLUMNS))}{summary_row}")
totals_cell = ws.cell(row=summary_row, column=8)
totals_cell.value = (
    "GP: 23,864  |  Robi: 18,838  |  BL BTS: 15,154  |  BTCL Latest: 24,073  |  "
    "Fiber Network: 27,259 (lines+pts)  |  BR Fiber: 707  |  "
    "BTCL Old: 30,379  |  Summit: 37,719  |  Fiber@Home: 113,358  |  "
    "Bahon: 20,580  |  BL Historical: 13,380"
)
totals_cell.font      = Font(name="Calibri", size=9, color="FFFFFF")
totals_cell.fill      = fill(C_TITLE_BG)
totals_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[summary_row].height = 28

# ── Freeze panes below header ─────────────────────────────────────────
ws.freeze_panes = "A5"

# ── Auto-filter on header row ─────────────────────────────────────────
ws.auto_filter.ref = f"A4:{get_column_letter(len(COLUMNS))}4"

# ── Page setup ────────────────────────────────────────────────────────
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToPage   = True
ws.page_setup.fitToWidth  = 1
ws.page_setup.fitToHeight = 0
ws.print_title_rows = "1:4"

wb.save(str(OUT))
print(f"Saved -> {OUT}")
